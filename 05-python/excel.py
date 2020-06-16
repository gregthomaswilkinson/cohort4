from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname

# def data_validation():
#     dvDecimal = DataValidation(type="decimal", formula1="{0}!$C$1:$C$10".format(quote_sheetname("product")), allow_blank=True)
#     print(dvDecimal)
    # dvDate = DataValidation(type="date", allow_blank=True)
    # dvString = DataValidation(type="text", allow_blank=True)
    # dvWholeNumber = DataValidation(type="whole", allow_blank=True)


def breakIntoDict():
    wb = load_workbook('excel.xlsx')
    builtDict = {}
    for sheet in wb.sheetnames:
        builtDict[sheet] = {}
        rowCount = 1
        for row in wb[sheet].iter_rows(min_row=2,min_col=1):
            if row[0].value is not None:
                col = 0
                builtDict[sheet][rowCount] = {}
                for cell in row:
                    builtDict[sheet][rowCount][wb[sheet][1][col].value] = cell.value
                    col += 1
            rowCount += 1
    # print(builtDict)
    return builtDict

# returns the customer ID(from invoice dictionary)
def getCustID(invoice) :
    for inv in xlDict['invoices']:
        if invoice == xlDict['invoices'][inv]['Invoice']:
            return xlDict['invoices'][inv]['Customer']

# returns the date (from invoice dictionary)
def getDate(invoice):
    for inv in xlDict['invoices']:
        if invoice == xlDict['invoices'][inv]['Invoice']:
            return xlDict['invoices'][inv]['Date']

# returns the customer name (from customers dictionary)
def getCustName(custID):
    for cust in xlDict['customers']:
        if custID == xlDict['customers'][cust]['ID']:
            return xlDict['customers'][cust]['Name']

# return the product ID and qty (from invoice_line_items)
# format [(product ID,qty),(p,q)...(p,q)]
def getInvLines(invoice):
    prodQtyPairs = []
    for invLine in xlDict['invoice_line_items']:
        if invoice == xlDict['invoice_line_items'][invLine]['Invoice']:
            prod = xlDict['invoice_line_items'][invLine]['Product']
            qty = xlDict['invoice_line_items'][invLine]['Quantity']
            prodQtyPairs.append(tuple((prod,qty)))
    return prodQtyPairs        

# returns product qty, name, price (from product dictionary)
# format {1: {'qty': #, 'product': 'AAA', 'price':##.##}}
#        {2: {'qty': #, 'product': 'AAA', 'price':##.##}}
def getProdPrice(prodQty):
    counter = 1
    invDict = {}
    for item in prodQty:
        for product in xlDict['product']:
            if item[0] == xlDict['product'][product]['ID']:
                invDict[counter] = {}
                ppName = xlDict['product'][product]['Name'] 
                ppPrice = xlDict['product'][product]['Price']
                ppQty = item[1]
                invDict[counter]['qty'] = ppQty
                invDict[counter]['product'] = ppName
                invDict[counter]['price'] = ppPrice
                counter=counter+1
    return invDict

# total price (calculated)

def genReport(invoiceReq):
    invReport = {}
    # Invoice requires:
    # Invoice number (from invoice dictionary)
    invReport['invoice'] = invoiceReq
    invReport['cust ID'] = getCustID(invoiceReq)
    invReport['date'] = getDate(invoiceReq)
    invReport['customer'] = getCustName(invReport['cust ID'])
    myInvLines = getInvLines(invoiceReq)
    # print(myInvLines)
    invReport['items'] = {}
    invReport['items'] = getProdPrice(myInvLines)
    return invReport

def printInvoice(invObj):
    invoiceFile = 'Invoice'+str(invObj['invoice'])
    with open(invoiceFile+'.txt','w') as invoice:
        # report.write(b[0]+'-'+b[1]+':'+str(tupleDict[b])+'\n')
        invoice.write('Invoice: '+str(invObj['invoice'])+'\n')
        invoice.write('Date: '+str(invObj['date'].date())+'\n')
        invoice.write('Customer: '+invObj['customer']+'\n')
        invoice.write('Customer ID: '+str(invObj['cust ID'])+'\n\n')
        invoice.write("{: <5} {: <31} {: <8}\n".format('Qty:','Product:','Price:'))
        invoice.write("{: <5} {: <31} {: <8}\n".format('====','========','======'))
        total = 0
        for row in invObj['items']:
            qty = str(invObj['items'][row]['qty'])
            product = invObj['items'][row]['product']
            total+=invObj['items'][row]['price']
            price = str(invObj['items'][row]['price'])
            invoice.write("{: <5} {: <31} {: <8}\n".format(qty,product,price))
        invoice.write("{: >44}\n".format('TOTAL:'))
        total = round(total,2)
        invoice.write("{: >44}\n".format(total))



xlDict = breakIntoDict()

# This block of code creates a list of existing invoice
# number to display to the user
invoices = xlDict['invoices']
invoiceList = []
for invoice in invoices:
    invoiceList.append(invoices[invoice]['Invoice'])
print(invoiceList)

# This code gets a valid invoice number from the user
invoiceReq = 0
if __name__ == "__main__":
    while invoiceReq not in invoiceList:
        invoiceReq = int(input('Enter existing invoice:'))
    invObj = genReport(invoiceReq)
    printInvoice(invObj)